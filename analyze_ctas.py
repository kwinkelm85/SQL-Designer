import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os

def apply_global_formatting(output_file):
    from openpyxl import load_workbook
    wb = load_workbook(output_file)
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    bold_font = Font(bold=True)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border
                    val_str = str(cell.value).strip()
                    if 'Only exists in CTAS' in val_str:
                        cell.fill = red_fill
                    elif 'Only exists in CTASEL' in val_str:
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    elif 'Shared' in val_str:
                        cell.fill = green_fill
                        
        for cell in ws[1]:
            cell.font = bold_font
            
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value is not None:
                        val_len = len(str(cell.value))
                        if val_len > max_length:
                            max_length = val_len
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 75)
            ws.column_dimensions[column_letter].width = adjusted_width
            
    wb.save(output_file)

def run_analysis():
    print("Loading CTAS Dataset...")
    source_file = 'DA_DESCRIBE CTAS.xlsx'
    xl = pd.ExcelFile(source_file)
    
    df_ctas = xl.parse('CTAS')
    df_ctasel = xl.parse('CTASEL')
    
    df_ctas.columns = df_ctas.columns.str.strip()
    df_ctasel.columns = df_ctasel.columns.str.strip()
    
    core_cols = ['DATABASE_NM', 'SCHEMA_NM', 'OBJECT_NM', 'COLUMN_NM', 'DATA_TYPE', 'DATA_LENGTH', 'DATA_PRECISION', 'DATA_SCALE', 'NULLABLE_YN', 'DATA_DEFAULT']
    
    for c in ['DATABASE_NM', 'SCHEMA_NM', 'OBJECT_NM', 'COLUMN_NM']:
        df_ctas[c] = df_ctas[c].astype(str).str.strip().str.upper()
        df_ctasel[c] = df_ctasel[c].astype(str).str.strip().str.upper()
        
    df_ctas = df_ctas[df_ctas['DATABASE_NM'] != 'NAN']
    df_ctasel = df_ctasel[df_ctasel['DATABASE_NM'] != 'NAN']
    
    df_ctas['ENV_SCHEMA'] = df_ctas['DATABASE_NM'] + "." + df_ctas['SCHEMA_NM']
    df_ctasel['ENV_SCHEMA'] = df_ctasel['DATABASE_NM'] + "." + df_ctasel['SCHEMA_NM']

    print("Deduplicating natively across Database granularities...")
    df_c = df_ctas.fillna('').groupby(['DATABASE_NM', 'SCHEMA_NM', 'ENV_SCHEMA', 'OBJECT_NM', 'COLUMN_NM'], as_index=False).first()
    df_e = df_ctasel.fillna('').groupby(['DATABASE_NM', 'SCHEMA_NM', 'ENV_SCHEMA', 'OBJECT_NM', 'COLUMN_NM'], as_index=False).first()
    
    databases = sorted(list(set(df_c['DATABASE_NM']).union(set(df_e['DATABASE_NM']))))
    
    summary_data = []
    
    all_missing_tables = []
    all_missing_columns = []
    all_type_mismatches = []
    
    for db in databases:
        db_c = df_c[df_c['DATABASE_NM'] == db]
        db_e = df_e[df_e['DATABASE_NM'] == db]
        
        # Identity cross-validation schema nomenclature natively derived per database layer
        c_schema_val = db_c['SCHEMA_NM'].unique()[0] if not db_c.empty else "N/A"
        e_schema_val = db_e['SCHEMA_NM'].unique()[0] if not db_e.empty else "N/A"
        
        cross_ref_string = f"{db}.{c_schema_val} vs {db}.{e_schema_val}"
        
        tbl_c = db_c[['DATABASE_NM', 'ENV_SCHEMA', 'OBJECT_NM']].drop_duplicates()
        tbl_e = db_e[['DATABASE_NM', 'ENV_SCHEMA', 'OBJECT_NM']].drop_duplicates()
        
        # Merge structurally on OBJECT_NM strictly contained internal to this identical DATABASE block
        tbl_m = pd.merge(tbl_c[['OBJECT_NM']], tbl_e[['OBJECT_NM']], on=['OBJECT_NM'], how='outer', indicator=True)
        
        shared_tables = tbl_m[tbl_m['_merge'] == 'both'].drop(columns=['_merge'])
        miss_ctasel_tbl = tbl_m[tbl_m['_merge'] == 'left_only'].drop(columns=['_merge'])
        miss_ctas_tbl = tbl_m[tbl_m['_merge'] == 'right_only'].drop(columns=['_merge'])
        
        miss_ctasel_tbl = pd.merge(miss_ctasel_tbl, tbl_c, on='OBJECT_NM', how='left')
        miss_ctas_tbl = pd.merge(miss_ctas_tbl, tbl_e, on='OBJECT_NM', how='left')
        
        if not miss_ctasel_tbl.empty:
            miss_ctasel_tbl['Missing Form'] = 'Only exists in CTAS'
            all_missing_tables.append(miss_ctasel_tbl)
        if not miss_ctas_tbl.empty:
            miss_ctas_tbl['Missing Form'] = 'Only exists in CTASEL'
            all_missing_tables.append(miss_ctas_tbl)
            
        # -- COLUMNS --
        col_c = pd.merge(db_c, shared_tables, on=['OBJECT_NM'], how='inner')
        col_e = pd.merge(db_e, shared_tables, on=['OBJECT_NM'], how='inner')
        
        col_m = pd.merge(col_c[['OBJECT_NM', 'COLUMN_NM']].drop_duplicates(), 
                         col_e[['OBJECT_NM', 'COLUMN_NM']].drop_duplicates(), 
                         on=['OBJECT_NM', 'COLUMN_NM'], how='outer', indicator=True)
                         
        shared_columns_df = col_m[col_m['_merge'] == 'both'].drop(columns=['_merge'])
        miss_csl_c = col_m[col_m['_merge'] == 'left_only'].drop(columns=['_merge'])
        miss_cts_c = col_m[col_m['_merge'] == 'right_only'].drop(columns=['_merge'])
        
        if not miss_csl_c.empty:
            miss_csl_c = pd.merge(miss_csl_c, db_c[['DATABASE_NM', 'ENV_SCHEMA', 'OBJECT_NM', 'COLUMN_NM']].drop_duplicates(), on=['OBJECT_NM', 'COLUMN_NM'], how='left')
            miss_csl_c['Missing Form'] = 'Only exists in CTAS'
            all_missing_columns.append(miss_csl_c)
            
        if not miss_cts_c.empty:
            miss_cts_c = pd.merge(miss_cts_c, db_e[['DATABASE_NM', 'ENV_SCHEMA', 'OBJECT_NM', 'COLUMN_NM']].drop_duplicates(), on=['OBJECT_NM', 'COLUMN_NM'], how='left')
            miss_cts_c['Missing Form'] = 'Only exists in CTASEL'
            all_missing_columns.append(miss_cts_c)
            
        # -- TYPE MISMATCHES --
        shared_full = pd.merge(col_c, col_e, on=['OBJECT_NM', 'COLUMN_NM'], suffixes=('_CTAS', '_CTASEL'))
        
        mismatches = []
        attributes = ['DATA_TYPE', 'DATA_LENGTH', 'DATA_PRECISION', 'DATA_SCALE', 'NULLABLE_YN', 'DATA_DEFAULT']
        
        for idx, row in shared_full.iterrows():
            diffs = []
            for prop in attributes:
                vc = str(row[f"{prop}_CTAS"])
                ve = str(row[f"{prop}_CTASEL"])
                if vc != ve:
                    diffs.append(f"{prop}: '{vc}' vs '{ve}'")
                    
            if diffs:
                mismatches.append({
                    'CROSS_REFERENCE': cross_ref_string,
                    'OBJECT_NM': row['OBJECT_NM'],
                    'COLUMN_NM': row['COLUMN_NM'],
                    'Mismatched Attributes': " | ".join(diffs)
                })
        
        if mismatches:
            all_type_mismatches.extend(mismatches)
            
        # -- BUILD SUMMARY --
        summary_data.append({
            'CROSS_REFERENCE (DATABASE_NM.SCHEMA_NM)': cross_ref_string,
            'Shared Tables': len(shared_tables),
            'Tables ONLY in CTAS': len(miss_ctasel_tbl),
            'Tables ONLY in CTASEL': len(miss_ctas_tbl),
            'Shared Columns': len(shared_columns_df),
            'Columns ONLY in CTAS': len(miss_csl_c),
            'Columns ONLY in CTASEL': len(miss_cts_c),
            'Column Type Mismatches': len(mismatches)
        })

    print("Compiling global output matrices...")
    df_summary = pd.DataFrame(summary_data)
    final_tables_df = pd.concat(all_missing_tables).sort_values(['DATABASE_NM', 'OBJECT_NM']) if all_missing_tables else pd.DataFrame()
    final_columns_df = pd.concat(all_missing_columns).sort_values(['DATABASE_NM', 'OBJECT_NM', 'COLUMN_NM']) if all_missing_columns else pd.DataFrame()
    final_mismatch_df = pd.DataFrame(all_type_mismatches) if all_type_mismatches else pd.DataFrame()

    output_file = 'DA_DESCRIBE_CTAS_Analysis.xlsx'
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        if not final_tables_df.empty:
            final_tables_df.to_excel(writer, sheet_name='Missing_Tables', index=False)
        if not final_columns_df.empty:
            final_columns_df.to_excel(writer, sheet_name='Missing_Columns', index=False)
        if not final_mismatch_df.empty:
            final_mismatch_df.to_excel(writer, sheet_name='Type_Mismatches', index=False)

    print("Injecting aesthetic formulas...")
    apply_global_formatting(output_file)
    print(f"DONE. Finished writing {output_file}")


if __name__ == "__main__":
    run_analysis()
