import pandas as pd
import numpy as np
import os
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

db_order = ['DEVJ', 'SIJ', 'TSTJ', 'JUSTIN', 'TRNJ', 'TRNE']

def analyze_sheet(df, object_type_name):
    print(f"[{object_type_name}] Analyzing Presence...")
    df.columns = df.columns.str.strip()
    df['DATABASE_NM'] = pd.Categorical(df['DATABASE_NM'], categories=db_order, ordered=True)

    # 1. Object Presence Matrix
    df_existence = df[['DATABASE_NM', 'OBJECT_NM']].drop_duplicates()
    presence_matrix = pd.pivot_table(df_existence, index='OBJECT_NM', columns='DATABASE_NM', aggfunc=len, fill_value=0)
    presence_matrix = presence_matrix.map(lambda x: 'Exists' if x > 0 else 'Missing')
    
    missing_objs = presence_matrix[presence_matrix.apply(lambda row: 'Missing' in row.values, axis=1)]

    print(f"[{object_type_name}] Analyzing Privilege Differences...")
    # 2. Privilege Differences
    df_privs = df[['DATABASE_NM', 'OBJECT_NM', 'GRANTEE', 'PRIVILEGE', 'GRANTABLE']].dropna(subset=['GRANTEE', 'PRIVILEGE'])
    
    df_privs['Rule_Exists'] = 'Yes'
    priv_matrix = pd.pivot_table(df_privs, 
                                 index=['OBJECT_NM', 'GRANTEE', 'PRIVILEGE', 'GRANTABLE'], 
                                 columns='DATABASE_NM', 
                                 values='Rule_Exists', 
                                 aggfunc='first', 
                                 fill_value='Missing')
    
    priv_diffs = priv_matrix[priv_matrix.apply(lambda row: 'Missing' in row.values, axis=1)]

    diffs_flat = priv_diffs.reset_index().melt(
        id_vars=['OBJECT_NM', 'GRANTEE', 'PRIVILEGE', 'GRANTABLE'],
        var_name='DATABASE_NM',
        value_name='Status'
    )
    
    diffs_flat['DATABASE_NM'] = pd.Categorical(diffs_flat['DATABASE_NM'], categories=db_order, ordered=True)
    diffs_flat = diffs_flat[['OBJECT_NM', 'DATABASE_NM', 'GRANTEE', 'PRIVILEGE', 'GRANTABLE', 'Status']]
    diffs_flat = diffs_flat.sort_values(by=['OBJECT_NM', 'DATABASE_NM', 'GRANTEE', 'PRIVILEGE'])
    diffs_grouped = diffs_flat.set_index(['OBJECT_NM', 'DATABASE_NM'])

    # 3. Child Item Analysis
    child_matrix = None
    child_mismatches = None
    if 'CHILD_ITEM_COUNT' in df.columns:
        df_children = df[['DATABASE_NM', 'OBJECT_NM', 'CHILD_ITEM_COUNT']].dropna().drop_duplicates()
        if not df_children.empty:
            print(f"[{object_type_name}] Analyzing Child Items...")
            child_matrix = pd.pivot_table(df_children, index='OBJECT_NM', columns='DATABASE_NM', values='CHILD_ITEM_COUNT', aggfunc='first')
            
            def is_mismatch(row):
                unique_counts = row.dropna().unique()
                return len(unique_counts) > 1
                
            child_mismatches = child_matrix[child_matrix.apply(is_mismatch, axis=1)]

    # 4. Extract Apparent Grantees
    df_grantees = None
    if not df_privs.empty:
        df_grantees = df_privs[['DATABASE_NM', 'GRANTEE']].drop_duplicates()

    return presence_matrix, missing_objs, diffs_grouped, child_matrix, child_mismatches, priv_matrix, df_grantees


def analyze():
    print("Loading data...")
    source_file = 'DA_DESCRIBE JUSTIN.xlsx'
    
    xl = pd.ExcelFile(source_file)
    results = {}
    
    overall_total_objects = 0
    overall_inconsistent_objects = 0
    db_metrics = {db: {'total_objects': 0, 'missing_objects': 0, 'missing_grants': 0} for db in db_order}
    
    obj_metrics = {}
    all_grantees_list = []
    
    for sheet_name in xl.sheet_names:
        if sheet_name in ['Users', 'Database Roles', 'Database Role Privs', 'APPL4 Roles', 'Roles', 'Role Privs']:
            continue
            
        df = xl.parse(sheet_name)
        res = analyze_sheet(df, sheet_name)
        results[sheet_name] = res
        
        presence_matrix = res[0]
        priv_matrix = res[5]
        df_grantees = res[6]
        
        if df_grantees is not None:
            all_grantees_list.append(df_grantees)
        
        sheet_tot_objs = len(presence_matrix)
        overall_total_objects += sheet_tot_objs
        
        inconsistent_mask = presence_matrix.apply(lambda row: 'Missing' in row.values, axis=1)
        overall_inconsistent_objects += inconsistent_mask.sum()
        
        obj_metrics[sheet_name] = {
            'total_objects': sheet_tot_objs,
            'db_metrics': {db: {'total_objects': 0, 'missing_objects': 0, 'missing_grants': 0} for db in db_order}
        }
        
        # Add to DB specific and Object specific metrics
        for db in db_order:
            if db in presence_matrix.columns:
                ex = (presence_matrix[db] == 'Exists').sum()
                ms = (presence_matrix[db] == 'Missing').sum()
            else:
                ex = 0
                ms = len(presence_matrix)
                
            db_metrics[db]['total_objects'] += ex
            db_metrics[db]['missing_objects'] += ms
            obj_metrics[sheet_name]['db_metrics'][db]['total_objects'] = ex
            obj_metrics[sheet_name]['db_metrics'][db]['missing_objects'] = ms
            
            if priv_matrix is not None and not priv_matrix.empty:
                if db in priv_matrix.columns:
                    gr = (priv_matrix[db] == 'Missing').sum()
                else:
                    gr = len(priv_matrix)
                    
                db_metrics[db]['missing_grants'] += gr
                obj_metrics[sheet_name]['db_metrics'][db]['missing_grants'] = gr

    print("Building Summary Dashboard...")
    app_consistency = (overall_inconsistent_objects / overall_total_objects) * 100 if overall_total_objects else 0
    total_obj_ddl = sum(m['missing_objects'] for m in db_metrics.values())
    total_grant_ddl = sum(m['missing_grants'] for m in db_metrics.values())
    
    summary_data = []
    summary_data.append(["Application Consistency (% of total objects not in all databases)", "", f"{app_consistency:.1f}%", "", ""])
    summary_data.append(["Overall Required Oracle DDL Statements (Object Creation)", "", total_obj_ddl, "", ""])
    summary_data.append(["Overall Required Oracle DDL Statements (Grant Creation)", "", total_grant_ddl, "", ""])
    summary_data.append(["", "", "", "", ""])
    summary_data.append(["DATABASE", "OBJECT TYPE", "Database to Application Consistency", "Required DDL Statements for Object Creation", "Required DDL Statements for Grant Creation"])
    
    for db in db_order:
        db_total = db_metrics[db]['total_objects']
        db_cons = (db_total / overall_total_objects) * 100 if overall_total_objects else 0
        
        summary_data.append([db, "OVERALL", f"{db_cons:.1f}%", db_metrics[db]['missing_objects'], db_metrics[db]['missing_grants']])
        
        for sheet_name in results.keys():
            stot = obj_metrics[sheet_name]['total_objects']
            s_db_tot = obj_metrics[sheet_name]['db_metrics'][db]['total_objects']
            s_db_cons = (s_db_tot / stot) * 100 if stot else 0
            
            s_db_obj_ddl = obj_metrics[sheet_name]['db_metrics'][db]['missing_objects']
            s_db_grant_ddl = obj_metrics[sheet_name]['db_metrics'][db]['missing_grants']
            
            summary_data.append(["", sheet_name, f"{s_db_cons:.1f}%", s_db_obj_ddl, s_db_grant_ddl])
        
    df_summary = pd.DataFrame(summary_data, columns=["C1", "C2", "C3", "C4", "C5"])
    
    print("Calculating Apparent Missing Grantees (JUSTIN vs TSTJ)...")
    if all_grantees_list:
        master_grantees = pd.concat(all_grantees_list).drop_duplicates()
        justin_grantees = set(master_grantees[master_grantees['DATABASE_NM'] == 'JUSTIN']['GRANTEE'])
        tstj_grantees_baseline = set(master_grantees[master_grantees['DATABASE_NM'] == 'TSTJ']['GRANTEE'])
        missing_in_tstj = sorted(list(justin_grantees - tstj_grantees_baseline))
    else:
        missing_in_tstj = []
        
    df_missing_grantees = pd.DataFrame({'Grantees apparently in JUSTIN but completely missing in TSTJ': missing_in_tstj})

    tstj_grantees_validation = set()

    for check_sheet, col_names in [
        ('Users', ['USERNAME']), 
        ('Roles', ['ROLE']), 
        ('Database Roles', ['ROLE']),
        ('Role Privs', ['GRANTEE', 'GRANTED_ROLE']),
        ('Database Role Privs', ['GRANTEE', 'GRANTED_ROLE'])
    ]:
        if check_sheet in xl.sheet_names:
            df_check = xl.parse(check_sheet)
            df_check.columns = df_check.columns.str.strip()
            
            # If the export includes a DATABASE column, filter cleanly.
            if 'DATABASE' in df_check.columns:
                df_check['DATABASE'] = df_check['DATABASE'].astype(str).str.strip()
                df_check = df_check[df_check['DATABASE'] == 'TSTJ']
                
            for col in col_names:
                if col in df_check.columns:
                    df_check[col] = df_check[col].astype(str).str.strip()
                    tstj_grantees_validation.update(df_check[col])
                    
    if len(tstj_grantees_validation) > 0:
        df_missing_grantees['Exists practically as a User/Role in TSTJ'] = df_missing_grantees.iloc[:, 0].apply(
            lambda x: "YES" if x in tstj_grantees_validation else "NO"
        )
        
    if 'APPL4 Roles' in xl.sheet_names:
        df_appl4 = xl.parse('APPL4 Roles')
        df_appl4.columns = df_appl4.columns.str.strip()
        
        if 'DATABASE' in df_appl4.columns:
            df_appl4['DATABASE'] = df_appl4['DATABASE'].astype(str).str.strip()
            df_appl4 = df_appl4[df_appl4['DATABASE'] == 'TSTJ']
            
        appl4_roles = set()
        if 'ROLE' in df_appl4.columns:
            df_appl4['ROLE'] = df_appl4['ROLE'].astype(str).str.strip()
            appl4_roles = set(df_appl4['ROLE'])
            
        if len(appl4_roles) > 0:
            df_missing_grantees['Exists in APPL4 (TSTJ)'] = df_missing_grantees.iloc[:, 0].apply(
                lambda x: "YES" if x in appl4_roles else "NO"
            )

    print("Saving to Excel...")
    output_file = 'DA_DESCRIBE_JUSTIN_Analysis_Summary.xlsx'
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        df_summary.to_excel(writer, sheet_name='Summary', index=False, header=False)
        worksheet_summary = writer.sheets['Summary']
        
        bold_font = Font(bold=True)
        for cell in ['A1', 'A2', 'A3', 'A4', 'B4', 'C4', 'D4', 'E4']:
            # Depending on if row 4 is blank or row 5 is headers, wait: I removed the blank row 4? No, row 4 is ["", "", "", ""]
            # Actually row 5 is ["DATABASE", "OBJECT TYPE"...] 
            worksheet_summary[cell].font = bold_font
            
        # We need headers to properly wrap now
        for cell in ['A5', 'B5', 'C5', 'D5', 'E5']:
            worksheet_summary[cell].font = bold_font
            worksheet_summary[cell].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            
        worksheet_summary.row_dimensions[5].height = 45
        
        # Merge the top title rows so they don't get cut off.
        worksheet_summary.merge_cells('A1:C1')
        worksheet_summary.merge_cells('A2:C2')
        worksheet_summary.merge_cells('A3:C3')
        
        worksheet_summary.column_dimensions['A'].width = 12
        worksheet_summary.column_dimensions['B'].width = 15
        worksheet_summary.column_dimensions['C'].width = 18
        worksheet_summary.column_dimensions['D'].width = 20
        worksheet_summary.column_dimensions['E'].width = 20
        
        # Write missing grantees sheet
        df_missing_grantees.to_excel(writer, sheet_name='Missing_Grantees_JUSTIN_TSTJ', index=False)
        worksheet_missing = writer.sheets['Missing_Grantees_JUSTIN_TSTJ']
        worksheet_missing.column_dimensions['A'].width = 75
        if len(df_missing_grantees.columns) > 1:
            worksheet_missing.column_dimensions['B'].width = 40
            worksheet_missing['B1'].font = bold_font
        if len(df_missing_grantees.columns) > 2:
            worksheet_missing.column_dimensions['C'].width = 25
            worksheet_missing['C1'].font = bold_font
            
        color_map = {
            'DEVJ': 'FFE5CC',
            'SIJ': 'FFCC99',
            'TSTJ': 'FFB266',
            'JUSTIN': 'FF9933',
            'TRNJ': 'FF8000',
            'TRNE': 'CC6600'
        }
        
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for sheet_name, (presence_matrix, missing_objs, diffs_grouped, child_matrix, child_mismatches, priv_matrix, df_grantees) in results.items():
            prefix = sheet_name[:12]
            
            presence_matrix.to_excel(writer, sheet_name=f'{prefix}_Overall_Matrix')
            missing_objs.to_excel(writer, sheet_name=f'{prefix}_Missing')
            
            mismatch_sheet_name = f'{prefix}_Priv_Mismatch'
            diffs_grouped.to_excel(writer, sheet_name=mismatch_sheet_name, index=True)

            workbook = writer.book
            if mismatch_sheet_name in writer.sheets:
                worksheet = writer.sheets[mismatch_sheet_name]
                max_row = len(diffs_grouped) + 1
                for db, color in color_map.items():
                    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    rule = CellIsRule(operator='equal', formula=[f'"{db}"'], fill=fill)
                    worksheet.conditional_formatting.add(f'B2:B{max_row}', rule)

            if child_matrix is not None and not child_matrix.empty:
                child_matrix.to_excel(writer, sheet_name=f'{prefix}_Child_Counts')
            
            if child_mismatches is not None and not child_mismatches.empty:
                child_mismatches.to_excel(writer, sheet_name=f'{prefix}_Child_Mismatch')

        # Apply universal formatting (borders, Yes/No colors, autofit columns)
        for ws in writer.book.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    val_str = str(cell.value).strip() if cell.value is not None else ""
                    if val_str in ("Exists", "YES"):
                        cell.fill = green_fill
                    elif val_str in ("Missing", "NO"):
                        cell.fill = red_fill
            
            if ws.title != 'Summary':
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value is not None:
                                val_len = len(str(cell.value))
                                if val_len > max_length:
                                    max_length = val_len
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 75)
                    # Keep any existing wider manual dimensions
                    if column_letter not in ws.column_dimensions or ws.column_dimensions[column_letter].width is None or ws.column_dimensions[column_letter].width < adjusted_width:
                        ws.column_dimensions[column_letter].width = adjusted_width

    print(f"Analysis complete! Saved to {output_file}")

if __name__ == '__main__':
    analyze()
